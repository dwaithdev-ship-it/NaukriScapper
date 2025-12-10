"""Data management for candidate profiles and exports."""
import os
import logging
from datetime import datetime
from typing import List, Dict, Optional
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

from models import get_session, JobSearch, Candidate, CallLog, init_db
from config import EXPORT_DIR

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class DataManager:
    """Manager for storing and exporting candidate data."""
    
    def __init__(self):
        """Initialize data manager."""
        init_db()
        self.session = get_session()
    
    def create_job_search(self, 
                         job_role: str,
                         experience_min: Optional[float] = None,
                         experience_max: Optional[float] = None,
                         location: Optional[str] = None,
                         job_type: Optional[str] = None) -> JobSearch:
        """
        Create a new job search record.
        
        Args:
            job_role: Job role/title
            experience_min: Minimum years of experience
            experience_max: Maximum years of experience
            location: Job location
            job_type: Type of job
            
        Returns:
            JobSearch object
        """
        job_search = JobSearch(
            job_role=job_role,
            experience_min=experience_min,
            experience_max=experience_max,
            location=location,
            job_type=job_type
        )
        self.session.add(job_search)
        self.session.commit()
        logger.info(f"Created job search: {job_search.id} - {job_role}")
        return job_search
    
    def add_candidate(self, job_search_id: int, profile_data: Dict) -> Candidate:
        """
        Add a candidate profile to the database.
        
        Args:
            job_search_id: ID of the related job search
            profile_data: Dictionary containing candidate information
            
        Returns:
            Candidate object
        """
        candidate = Candidate(
            job_search_id=job_search_id,
            name=profile_data.get('name'),
            email=profile_data.get('email'),
            phone=profile_data.get('phone'),
            current_location=profile_data.get('current_location'),
            experience_years=profile_data.get('experience_years'),
            current_company=profile_data.get('current_company'),
            current_designation=profile_data.get('current_designation'),
            skills=profile_data.get('skills'),
            education=profile_data.get('education'),
            profile_url=profile_data.get('profile_url'),
            resume_url=profile_data.get('resume_url'),
            notice_period=profile_data.get('notice_period'),
            current_salary=profile_data.get('current_salary'),
            expected_salary=profile_data.get('expected_salary'),
            comments=profile_data.get('comments', '')
        )
        self.session.add(candidate)
        self.session.commit()
        logger.info(f"Added candidate: {candidate.id} - {candidate.name}")
        return candidate
    
    def add_candidates_bulk(self, job_search_id: int, profiles: List[Dict]) -> List[Candidate]:
        """
        Add multiple candidates at once.
        
        Args:
            job_search_id: ID of the related job search
            profiles: List of profile dictionaries
            
        Returns:
            List of Candidate objects
        """
        candidates = []
        for profile in profiles:
            try:
                candidate = self.add_candidate(job_search_id, profile)
                candidates.append(candidate)
            except Exception as e:
                logger.error(f"Error adding candidate {profile.get('name')}: {e}")
        
        logger.info(f"Added {len(candidates)} candidates to job search {job_search_id}")
        return candidates
    
    def get_candidates_by_job_search(self, job_search_id: int) -> List[Candidate]:
        """Get all candidates for a specific job search."""
        return self.session.query(Candidate).filter(
            Candidate.job_search_id == job_search_id
        ).all()
    
    def get_all_candidates(self) -> List[Candidate]:
        """Get all candidates from database."""
        return self.session.query(Candidate).all()
    
    def update_candidate_status(self,
                               candidate_id: int,
                               contacted: Optional[bool] = None,
                               interested: Optional[bool] = None,
                               interview_scheduled: Optional[bool] = None,
                               interview_date: Optional[datetime] = None,
                               comments: Optional[str] = None):
        """
        Update candidate contact status.
        
        Args:
            candidate_id: Candidate ID
            contacted: Whether candidate has been contacted
            interested: Whether candidate is interested
            interview_scheduled: Whether interview is scheduled
            interview_date: Date/time of interview
            comments: Additional comments
        """
        candidate = self.session.query(Candidate).get(candidate_id)
        if not candidate:
            logger.error(f"Candidate {candidate_id} not found")
            return
        
        if contacted is not None:
            candidate.contacted = contacted
        if interested is not None:
            candidate.interested = interested
        if interview_scheduled is not None:
            candidate.interview_scheduled = interview_scheduled
        if interview_date is not None:
            candidate.interview_date = interview_date
        if comments is not None:
            candidate.comments = comments
        
        self.session.commit()
        logger.info(f"Updated candidate {candidate_id} status")
    
    def log_call(self,
                candidate_id: int,
                script_used: str,
                call_status: str,
                interested: Optional[bool] = None,
                ai_tool_used: Optional[str] = None,
                notes: Optional[str] = None) -> CallLog:
        """
        Log an AI call interaction.
        
        Args:
            candidate_id: Candidate ID
            script_used: Script used for the call
            call_status: Status of the call
            interested: Whether candidate showed interest
            ai_tool_used: Which AI tool was used (n8n, make.com, etc.)
            notes: Additional notes
            
        Returns:
            CallLog object
        """
        call_log = CallLog(
            candidate_id=candidate_id,
            script_used=script_used,
            call_status=call_status,
            interested=interested,
            ai_tool_used=ai_tool_used,
            notes=notes
        )
        self.session.add(call_log)
        
        # Update candidate status if interested is provided
        if interested is not None:
            candidate = self.session.query(Candidate).get(candidate_id)
            if candidate:
                candidate.contacted = True
                candidate.interested = interested
        
        self.session.commit()
        logger.info(f"Logged call for candidate {candidate_id}")
        return call_log
    
    def export_to_excel(self, 
                       job_search_id: Optional[int] = None,
                       filename: Optional[str] = None) -> str:
        """
        Export candidates to Excel spreadsheet.
        
        Args:
            job_search_id: Optional job search ID to filter by
            filename: Optional custom filename
            
        Returns:
            Path to the exported file
        """
        # Get candidates
        if job_search_id:
            candidates = self.get_candidates_by_job_search(job_search_id)
            job_search = self.session.query(JobSearch).get(job_search_id)
            default_filename = f"candidates_{job_search.job_role.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        else:
            candidates = self.get_all_candidates()
            default_filename = f"all_candidates_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        if not candidates:
            logger.warning("No candidates to export")
            return ""
        
        # Prepare filename
        filename = filename or default_filename
        filepath = EXPORT_DIR / filename
        
        # Create DataFrame
        data = []
        for c in candidates:
            data.append({
                'ID': c.id,
                'Name': c.name,
                'Email': c.email,
                'Phone': c.phone,
                'Location': c.current_location,
                'Experience (Years)': c.experience_years,
                'Current Company': c.current_company,
                'Current Designation': c.current_designation,
                'Skills': c.skills,
                'Education': c.education,
                'Profile URL': c.profile_url,
                'Notice Period': c.notice_period,
                'Current Salary': c.current_salary,
                'Expected Salary': c.expected_salary,
                'Contacted': 'Yes' if c.contacted else 'No',
                'Interested': 'Yes' if c.interested else ('No' if c.interested is False else 'Unknown'),
                'Interview Scheduled': 'Yes' if c.interview_scheduled else 'No',
                'Interview Date': c.interview_date.strftime('%Y-%m-%d %H:%M') if c.interview_date else '',
                'Comments': c.comments or '',
                'Scraped Date': c.scraped_date.strftime('%Y-%m-%d %H:%M'),
            })
        
        df = pd.DataFrame(data)
        
        # Create Excel workbook with formatting
        wb = Workbook()
        ws = wb.active
        ws.title = "Candidates"
        
        # Add headers with formatting
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:  # Header row
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save workbook
        wb.save(filepath)
        logger.info(f"Exported {len(candidates)} candidates to {filepath}")
        return str(filepath)
    
    def get_statistics(self, job_search_id: Optional[int] = None) -> Dict:
        """
        Get statistics about candidates.
        
        Args:
            job_search_id: Optional job search ID to filter by
            
        Returns:
            Dictionary with statistics
        """
        query = self.session.query(Candidate)
        if job_search_id:
            query = query.filter(Candidate.job_search_id == job_search_id)
        
        all_candidates = query.all()
        
        stats = {
            'total_candidates': len(all_candidates),
            'contacted': sum(1 for c in all_candidates if c.contacted),
            'interested': sum(1 for c in all_candidates if c.interested),
            'not_interested': sum(1 for c in all_candidates if c.interested is False),
            'interview_scheduled': sum(1 for c in all_candidates if c.interview_scheduled),
            'pending_contact': sum(1 for c in all_candidates if not c.contacted),
        }
        
        return stats
    
    def close(self):
        """Close database session."""
        self.session.close()


if __name__ == '__main__':
    # Example usage
    dm = DataManager()
    
    # Create a job search
    job_search = dm.create_job_search(
        job_role="Python Developer",
        experience_min=2.0,
        experience_max=5.0,
        location="Bangalore",
        job_type="Full-time"
    )
    
    # Add sample candidates
    sample_profiles = [
        {
            'name': 'John Doe',
            'email': 'john.doe@example.com',
            'phone': '+91-9876543210',
            'current_location': 'Bangalore',
            'experience_years': 3.5,
            'current_company': 'Tech Corp',
            'current_designation': 'Software Engineer',
            'skills': 'Python, Django, REST API, PostgreSQL',
            'education': 'B.Tech in Computer Science'
        },
        {
            'name': 'Jane Smith',
            'email': 'jane.smith@example.com',
            'phone': '+91-9876543211',
            'current_location': 'Bangalore',
            'experience_years': 4.0,
            'current_company': 'Software Solutions',
            'current_designation': 'Senior Developer',
            'skills': 'Python, Flask, AWS, Docker',
            'education': 'M.Tech in Computer Science'
        }
    ]
    
    dm.add_candidates_bulk(job_search.id, sample_profiles)
    
    # Export to Excel
    excel_file = dm.export_to_excel(job_search.id)
    print(f"Exported to: {excel_file}")
    
    # Get statistics
    stats = dm.get_statistics(job_search.id)
    print(f"Statistics: {stats}")
    
    dm.close()
